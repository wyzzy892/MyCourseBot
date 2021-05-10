import openpyxl
import random
import time, xlrd


def convert_to_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    for id, d in enumerate(data):
        for col, item in enumerate(d):
            ws.cell(column=col+1, row=id+1, value=f'{item}')
    filename = f"temp_{random.randint(0, 10000000)}_{time.time()}.xlsx"
    wb.save(filename)
    return filename


def convert_to_list(filename):
    wb = xlrd.open_workbook(file_contents=filename)
    ws = wb.sheet_by_index(0)
    list_report = []
    for row_x in range(0, ws.nrows):
        list_row = []
        for col_x in range(0, ws.ncols):
            list_row.append(ws.cell(row_x, col_x).value)
        list_report.append(list_row)
    return list_report

